import streamlit as st
import fitz  # PyMuPDF
import docx2txt
import pandas as pd
import io
import shutil
from rapidfuzz import fuzz
from openpyxl import load_workbook
import requests

st.set_page_config(page_title="AI-сервис подбора", layout="wide")
st.title("🤖 AI-сервис подбора оборудования")
st.markdown("Загрузите техническое задание и прайс-листы — система всё сделает сама.")

# Блок загрузки ТЗ
st.header("📄 Техническое задание")
uploaded_spec = st.file_uploader("Загрузите файл с ТЗ (PDF, DOCX)", type=["pdf", "docx"])

# Блок загрузки прайсов
st.header("📊 Прайсы поставщиков")
uploaded_prices = st.file_uploader("Загрузите 1 или несколько прайсов (Excel)", type=["xlsx"], accept_multiple_files=True)

# Блок загрузки скидок
st.header("💸 Скидки от поставщиков (по желанию)")
discounts_file = st.file_uploader("Файл со скидками (Excel)", type=["xlsx"], accept_multiple_files=False)

# OCR API (через ocr.space)
def extract_text_ocr_api(file_bytes):
    url = 'https://api.ocr.space/parse/image'
    response = requests.post(
        url,
        files={"file": file_bytes},
        data={
            "apikey": "K86918490388957",
            "language": "rus",
            "isOverlayRequired": False,
        },
    )
    try:
        result = response.json()
        return result["ParsedResults"][0]["ParsedText"]
    except:
        return ""

# Функция обработки ТЗ
def extract_text_from_spec(file):
    if file.name.endswith(".pdf"):
        try:
            pdf_doc = fitz.open(stream=file.read(), filetype="pdf")
            text = "".join([page.get_text() for page in pdf_doc])
            if len(text.strip()) < 10:
                file.seek(0)
                return extract_text_ocr_api(file)
            return text
        except:
            file.seek(0)
            return extract_text_ocr_api(file)
    elif file.name.endswith(".docx"):
        return docx2txt.process(file)
    return ""

# Функция чтения прайсов с автоматическим поиском заголовков
def read_prices(files):
    dfs = []
    for f in files:
        try:
            df_raw = pd.read_excel(f, header=None)
            header_row = df_raw.apply(lambda x: x.astype(str).str.contains("артикул|наименование|номенклатура", case=False).any(), axis=1)
            if header_row.any():
                idx = header_row.idxmax()
                df = pd.read_excel(f, skiprows=idx)
                df['Поставщик'] = f.name
                dfs.append(df)
        except:
            st.warning(f"Не удалось прочитать файл: {f.name}")
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

# Загрузка скидок

def load_discounts(discount_file):
    if discount_file:
        df = pd.read_excel(discount_file)
        return dict(zip(df.iloc[:, 0].astype(str), df.iloc[:, 1]))
    return {}

# Поиск лучших совпадений

def match_top_variants(spec_text, df_prices, top_n=3):
    results = []
    for line in spec_text.split("\n"):
        line = line.strip()
        if len(line) < 5:
            continue
        candidates = []
        for _, row in df_prices.iterrows():
            row_str = " ".join([str(v) for v in row.values if isinstance(v, str)])
            score = fuzz.token_sort_ratio(line.lower(), row_str.lower())
            candidates.append((score, row))
        top_matches = sorted(candidates, key=lambda x: x[0], reverse=True)[:top_n]
        for score, row in top_matches:
            matched = row.to_dict()
            matched['Совпадение'] = score
            matched['Из ТЗ'] = line
            results.append(matched)
    return pd.DataFrame(results)

# Генерация Excel

def generate_template_excel(df_result, discounts):
    template_path = "Форма для результата.xlsx"
    export_path = "Готовый_результат_по_шаблону.xlsx"
    shutil.copy(template_path, export_path)
    wb = load_workbook(export_path)
    ws = wb.active
    start_row = 4
    for i, row in df_result.iterrows():
        r = start_row + i
        price = row.get("Цена", 0)
        qty = row.get("Количество", 1)
        supplier = row.get("Поставщик", "")
        discount = discounts.get(str(supplier), 0)
        ws[f"A{r}"] = i + 1
        ws[f"C{r}"] = row.get("Из ТЗ", "")
        ws[f"D{r}"] = row.get("Наименование", row.get("Аналог", ""))
        ws[f"E{r}"] = f"{row.get('Совпадение', '')}%"
        ws[f"F{r}"] = qty
        ws[f"G{r}"] = price
        ws[f"H{r}"] = f"=G{r}*F{r}"
        ws[f"K{r}"] = row.get("Ссылка", "")
        ws[f"M{r}"] = price
        ws[f"N{r}"] = supplier
        ws[f"O{r}"] = discount
        ws[f"P{r}"] = f"=H{r}*(1 - O{r}/100)"
    wb.save(export_path)
    return export_path

# Запуск анализа
if st.button("🚀 Запустить подбор"):
    if uploaded_spec and uploaded_prices:
        st.success("Файлы получены! Идёт обработка...")

        spec_text = extract_text_from_spec(uploaded_spec)
        if not spec_text.strip():
            st.error("Не удалось распознать текст из ТЗ. Проверьте формат файла.")
        st.subheader("📜 Распознанный текст из ТЗ")
        st.text_area("", spec_text, height=300)

        df_prices = read_prices(uploaded_prices)
        st.subheader("📋 Объединённый прайс-лист")
        st.dataframe(df_prices.astype(str).head(20))

        if not df_prices.empty and spec_text:
            df_result = match_top_variants(spec_text, df_prices, top_n=3)
            st.subheader("✅ Сопоставленные позиции (до 3 вариантов на каждую)")

            if not df_result.empty:
                min_match = st.slider("Минимальный процент совпадения", 0, 100, 70)
                keyword = st.text_input("🔍 Поиск по ключевому слову")

                filtered_df = df_result[df_result['Совпадение'] >= min_match]
                if keyword:
                    filtered_df = filtered_df[filtered_df.apply(lambda row: row.astype(str).str.contains(keyword, case=False).any(), axis=1)]

                st.dataframe(filtered_df.astype(str))

                discounts = load_discounts(discounts_file)

                if st.button("📄 Сформировать Excel по шаблону"):
                    file_path = generate_template_excel(filtered_df, discounts)
                    with open(file_path, "rb") as f:
                        st.download_button("📥 Скачать файл по шаблону", data=f, file_name="Готовый_результат_по_шаблону.xlsx")
            else:
                st.warning("Ничего не найдено для отображения.")
    else:
        st.warning("Пожалуйста, загрузите и ТЗ, и хотя бы один прайс.")
